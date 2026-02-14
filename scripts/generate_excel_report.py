"""
PayrollPro - Excel Report Generator
Generates professional Excel reports for payroll and employee data
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create output directory
output_dir = "/home/z/my-project/download"
os.makedirs(output_dir, exist_ok=True)

# Create workbook
wb = Workbook()

# Style definitions
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
data_font = Font(name='Times New Roman', size=11)
title_font = Font(name='Times New Roman', size=18, bold=True, color='000000')
alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# ==========================================
# Sheet 1: Employee Summary
# ==========================================
ws1 = wb.active
ws1.title = "Employee Summary"

ws1['B2'] = "PayrollPro - Employee Summary Report"
ws1['B2'].font = title_font
ws1['B2'].alignment = left_align

ws1['B3'] = "Report Date: November 2024"
ws1['B3'].font = Font(name='Times New Roman', size=11, color='666666')

headers = ['Branch', 'Total Employees', 'Present Today', 'On Leave', 'Late', 'Attendance Rate']
for col, header in enumerate(headers, 2):
    cell = ws1.cell(row=5, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

branch_data = [
    ['Algiers HQ', 145, 132, 8, 5, '91.0%'],
    ['Oran Branch', 78, 71, 4, 3, '91.0%'],
    ['Setif Branch', 52, 48, 2, 2, '92.3%'],
    ['Annaba Branch', 35, 31, 2, 2, '88.6%'],
    ['Tamanrasset', 25, 23, 1, 1, '92.0%'],
    ['Tindouf Branch', 18, 17, 1, 0, '94.4%'],
]

for row_idx, row_data in enumerate(branch_data, 6):
    for col_idx, value in enumerate(row_data, 2):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = center_align
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# Total row
totals = ['TOTAL', 353, 298, 18, 25, '84.4%']
for col_idx, value in enumerate(totals, 2):
    cell = ws1.cell(row=12, column=col_idx, value=value)
    cell.font = Font(name='Times New Roman', size=11, bold=True)
    cell.alignment = center_align

# Adjust column widths
widths1 = [5, 18, 18, 16, 12, 10, 16]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ==========================================
# Sheet 2: Payroll Summary
# ==========================================
ws2 = wb.create_sheet("Payroll Summary")

ws2['B2'] = "PayrollPro - Monthly Payroll Summary"
ws2['B2'].font = title_font
ws2['B2'].alignment = left_align

payroll_headers = ['Month', 'Total Payroll (M DZD)', 'Employees', 'Avg Salary (K DZD)', 'Growth']
for col, header in enumerate(payroll_headers, 2):
    cell = ws2.cell(row=5, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

payroll_data = [
    ['January', 27.2, 342, 78.0, '+2.1%'],
    ['February', 26.8, 345, 76.5, '-1.5%'],
    ['March', 27.5, 348, 79.2, '+2.6%'],
    ['April', 28.1, 352, 81.0, '+2.2%'],
    ['May', 27.9, 350, 78.5, '-0.7%'],
    ['June', 28.3, 353, 82.0, '+1.4%'],
    ['July', 28.0, 351, 79.8, '-1.1%'],
    ['August', 27.6, 349, 77.5, '-1.4%'],
    ['September', 28.2, 352, 81.5, '+2.2%'],
    ['October', 27.8, 350, 79.0, '-1.4%'],
    ['November', 28.5, 353, 82.5, '+2.5%'],
]

for row_idx, row_data in enumerate(payroll_data, 6):
    for col_idx, value in enumerate(row_data, 2):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = center_align
        if row_idx % 2 == 0:
            cell.fill = alt_fill
        if col_idx == 6 and isinstance(value, str):
            if value.startswith('+'):
                cell.font = Font(name='Times New Roman', size=11, color='008000')
            elif value.startswith('-'):
                cell.font = Font(name='Times New Roman', size=11, color='FF0000')

widths2 = [5, 14, 22, 12, 20, 10]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ==========================================
# Sheet 3: Department Distribution
# ==========================================
ws3 = wb.create_sheet("Department Distribution")

ws3['B2'] = "PayrollPro - Department Distribution"
ws3['B2'].font = title_font
ws3['B2'].alignment = left_align

dept_headers = ['Department', 'Employees', 'Percentage', 'Avg Salary (K DZD)']
for col, header in enumerate(dept_headers, 2):
    cell = ws3.cell(row=5, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

dept_data = [
    ['Engineering', 78, '22.1%', 95.0],
    ['Operations', 65, '18.4%', 68.0],
    ['Sales', 52, '14.7%', 72.0],
    ['Human Resources', 38, '10.8%', 65.0],
    ['Finance', 35, '9.9%', 85.0],
    ['Marketing', 32, '9.1%', 70.0],
    ['Administration', 28, '7.9%', 55.0],
    ['IT Support', 25, '7.1%', 75.0],
]

for row_idx, row_data in enumerate(dept_data, 6):
    for col_idx, value in enumerate(row_data, 2):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = center_align
        if row_idx % 2 == 0:
            cell.fill = alt_fill

widths3 = [5, 18, 12, 14, 20]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ==========================================
# Save
# ==========================================
output_path = os.path.join(output_dir, "PayrollPro_Report_November_2024.xlsx")
wb.save(output_path)
print(f"✅ Excel report saved: {output_path}")
